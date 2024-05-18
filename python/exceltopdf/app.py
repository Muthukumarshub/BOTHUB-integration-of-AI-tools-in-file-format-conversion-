from flask import Flask, render_template, request, send_file, jsonify
import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

app = Flask(__name__)

def convert_excel_to_pdf(excel_file_path):
    try:
        pdf_file_path = excel_file_path[:-5] + '.pdf'  # Change .xlsx to .pdf
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        c = canvas.Canvas(pdf_file_path, pagesize=letter)
        width, height = letter

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                text = str(cell_value) if cell_value is not None else ""
                c.drawString(72 + (col_idx - 1) * 100, height - 72 - (row_idx - 1) * 20, text)

        c.save()
        return pdf_file_path, None
    except Exception as e:
        return None, str(e)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file.filename.endswith('.pdf'):
        pdf_file_path = os.path.join('uploads', file.filename)
        file.save(pdf_file_path)
        ppt_file_path, error = convert_pdf_to_ppt(pdf_file_path)
        if error:
            return jsonify({'error': error}), 500
        return jsonify({
            'download_link': f'/download/{os.path.basename(ppt_file_path)}'
        })
    elif file.filename.endswith('.xlsx'):
        excel_file_path = os.path.join('uploads', file.filename)
        file.save(excel_file_path)
        pdf_file_path, error = convert_excel_to_pdf(excel_file_path)
        if error:
            return jsonify({'error': error}), 500
        return jsonify({
            'download_link': f'/download/{os.path.basename(pdf_file_path)}'
        })
    else:
        return jsonify({'error': 'Please upload a PDF or Excel file'}), 400

@app.route('/download/<filename>')
def download(filename):
    file_path = os.path.join('uploads', filename)
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(debug=True, port=5005)  # Set the port to 5004
